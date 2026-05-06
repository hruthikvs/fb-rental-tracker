#!/usr/bin/env python3
"""
Excel Generator for FB Rentals
--export-feedback : reads listings.xlsx → writes existing_data.json (preserves Status/Notes/Feedback)
--generate        : reads analyzed_posts.json + existing_data.json → writes listings.xlsx
"""

import json
import argparse
import sys
from pathlib import Path
from datetime import datetime

COLUMNS = [
    "Match Score",
    "Post Link",
    "Posted At",
    "Poster Name",
    "Society / Building",
    "Area / Locality",
    "Commute Estimate",
    "Flat Type",
    "Room Type",
    "Rent (₹)",
    "Maintenance (₹)",
    "Total Monthly (₹)",
    "Deposit (months)",
    "Brokerage",
    "Available From",
    "Furnishing",
    "Gated Community",
    "Key Amenities",
    "Contact Number",
    "Google Maps Link",
    "Veg Only Warning",
    "Bachelor Friendly",
    "Broker Flag",
    "Broker Reason",
    "Status",
    "Notes",
    "Feedback",
]

FIELD_MAP = {
    "match_score":       "Match Score",
    "url":               "Post Link",
    "timestamp_iso":     "Posted At",
    "author":            "Poster Name",
    "society_name":      "Society / Building",
    "area_locality":     "Area / Locality",
    "commute_estimate":  "Commute Estimate",
    "flat_type":         "Flat Type",
    "room_type":         "Room Type",
    "rent":              "Rent (₹)",
    "maintenance":       "Maintenance (₹)",
    "total_monthly":     "Total Monthly (₹)",
    "deposit_months":    "Deposit (months)",
    "brokerage":         "Brokerage",
    "available_from":    "Available From",
    "furnishing":        "Furnishing",
    "gated_community":   "Gated Community",
    "amenities_summary": "Key Amenities",
    "contact_number":    "Contact Number",
    "google_maps_link":  "Google Maps Link",
    "veg_only":          "Veg Only Warning",
    "bachelor_friendly": "Bachelor Friendly",
    "broker_flag":       "Broker Flag",
    "broker_reason":     "Broker Reason",
}

COL_WIDTHS = {
    "Match Score": 9,
    "Post Link": 11,
    "Posted At": 16,
    "Poster Name": 20,
    "Society / Building": 24,
    "Area / Locality": 16,
    "Commute Estimate": 16,
    "Flat Type": 9,
    "Room Type": 15,
    "Rent (₹)": 10,
    "Maintenance (₹)": 13,
    "Total Monthly (₹)": 15,
    "Deposit (months)": 13,
    "Brokerage": 13,
    "Available From": 15,
    "Furnishing": 15,
    "Gated Community": 13,
    "Key Amenities": 35,
    "Contact Number": 15,
    "Google Maps Link": 11,
    "Veg Only Warning": 12,
    "Bachelor Friendly": 14,
    "Broker Flag": 11,
    "Broker Reason": 28,
    "Status": 13,
    "Notes": 28,
    "Feedback": 28,
}

BROKER_FILLS = {
    "Likely":    "FFCCCC",
    "Suspected": "FFE0CC",
    "Doubt":     "FFF9C4",
}


# ── Export mode ──────────────────────────────────────────────────────────────

def export_feedback(excel_path: Path, out_path: Path):
    """Read listings.xlsx and export Status/Notes/Feedback to existing_data.json."""
    if not excel_path.exists():
        print(f"No existing Excel at {excel_path} — nothing to export.")
        out_path.write_text("[]", encoding="utf-8")
        return

    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]

    def col(name):
        return headers.index(name) if name in headers else None

    url_col      = col("Post Link")
    status_col   = col("Status")
    notes_col    = col("Notes")
    feedback_col = col("Feedback")

    if url_col is None:
        print("Warning: 'Post Link' column not found in existing Excel.")
        out_path.write_text("[]", encoding="utf-8")
        return

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[url_col]
        if not url:
            continue
        # Hyperlinked cells store the display value ("View Post"), not the URL.
        # We need the actual href — try reading the cell object.
        rows.append({
            "url":      str(url),
            "status":   str(row[status_col])   if status_col   is not None and row[status_col]   else "",
            "notes":    str(row[notes_col])     if notes_col    is not None and row[notes_col]    else "",
            "feedback": str(row[feedback_col])  if feedback_col is not None and row[feedback_col] else "",
        })

    # Re-read with data_only=False to get hyperlink URLs for "Post Link" cells
    wb2 = openpyxl.load_workbook(excel_path)
    ws2 = wb2.active
    url_col_letter_idx = url_col  # 0-based
    for i, row_obj in enumerate(ws2.iter_rows(min_row=2), 0):
        if i >= len(rows):
            break
        cell = row_obj[url_col_letter_idx]
        if cell.hyperlink:
            rows[i]["url"] = str(cell.hyperlink.target or cell.hyperlink)

    out_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Exported {len(rows)} existing rows to {out_path}")


# ── Generate mode ─────────────────────────────────────────────────────────────

def generate(analyzed_path: Path, existing_path: Path, output_path: Path):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    with open(analyzed_path, encoding="utf-8") as f:
        posts = json.load(f)

    # Build existing-data lookup keyed by URL
    existing = {}
    if existing_path.exists():
        with open(existing_path, encoding="utf-8") as f:
            for row in json.load(f):
                existing[row["url"]] = row

    # Sort by match_score descending
    posts.sort(key=lambda p: p.get("match_score", 0), reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    # ── Status dropdown ───────────────────────────────────────────────────────
    status_col_idx = COLUMNS.index("Status") + 1
    dv = DataValidation(
        type="list",
        formula1='"Visited,Interested,Rejected,Shortlisted"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv)

    # ── Data rows ─────────────────────────────────────────────────────────────
    feedback_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    link_font     = Font(color="0563C1", underline="single", name="Calibri", size=10)
    normal_font   = Font(name="Calibri", size=10)
    wrap_align    = Alignment(wrap_text=True, vertical="top")

    for row_idx, post in enumerate(posts, 2):
        url         = post.get("url", "")
        prev        = existing.get(url, {})
        broker_flag = post.get("broker_flag", "None") or "None"
        row_fill    = BROKER_FILLS.get(broker_flag)

        # Build column-name → value map
        values = {}
        for field, col_name in FIELD_MAP.items():
            values[col_name] = post.get(field, "")

        # Format timestamp
        ts = values.get("Posted At", "")
        if isinstance(ts, str) and "T" in ts:
            try:
                dt = datetime.fromisoformat(ts)
                values["Posted At"] = dt.strftime("%Y-%m-%d %H:%M")
            except Exception:
                pass

        # Preserved user columns (never overwrite)
        values["Status"]   = prev.get("status", "")   or ""
        values["Notes"]    = prev.get("notes", "")     or ""
        values["Feedback"] = prev.get("feedback", "")  or ""

        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=values.get(col_name, ""))
            cell.alignment = wrap_align

            if col_name == "Post Link" and url:
                cell.hyperlink = url
                cell.value     = "View Post"
                cell.font      = link_font
            elif col_name == "Google Maps Link" and values.get("Google Maps Link"):
                cell.hyperlink = values["Google Maps Link"]
                cell.value     = "Map"
                cell.font      = link_font
            elif col_name == "Feedback":
                cell.fill = feedback_fill
                cell.font = normal_font
                if row_fill:
                    # Feedback column keeps its blue even on broker rows
                    pass
            else:
                cell.font = normal_font
                if row_fill:
                    cell.fill = PatternFill(
                        start_color=row_fill, end_color=row_fill, fill_type="solid"
                    )

            if col_name == "Status":
                dv.add(cell)

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)

    wb.save(output_path)
    print(f"Excel saved: {output_path}  ({len(posts)} listings)")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Excel generator for FB rentals")
    group  = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--export-feedback", action="store_true",
                       help="Export Status/Notes/Feedback from existing Excel to existing_data.json")
    group.add_argument("--generate", action="store_true",
                       help="Generate listings.xlsx from analyzed_posts.json")

    parser.add_argument("--excel",    default="listings.xlsx",       help="Path to listings.xlsx")
    parser.add_argument("--analyzed", default="analyzed_posts.json", help="Path to analyzed_posts.json")
    parser.add_argument("--existing", default="existing_data.json",  help="Path to existing_data.json")
    args = parser.parse_args()

    if args.export_feedback:
        export_feedback(Path(args.excel), Path(args.existing))

    elif args.generate:
        analyzed_path = Path(args.analyzed)
        if not analyzed_path.exists():
            print(f"ERROR: {analyzed_path} not found. Cannot generate Excel.")
            return 1
        generate(analyzed_path, Path(args.existing), Path(args.excel))

    return 0


if __name__ == "__main__":
    sys.exit(main())
